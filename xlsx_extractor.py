import openpyxl
from util import extractor

def extract_data_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    excel_data = {}
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_data = []
        for row in sheet.iter_rows(values_only=True):
            sheet_data.append(list(row))
        excel_data[sheet_name] = sheet_data
    
    return excel_data

if __name__ == "__main__":
    input_excel = 'test_documents/test.xlsx'
    output_json = 'output_xlsx.json'
    extractor(input_excel, output_json, extract_data_from_excel, "file")